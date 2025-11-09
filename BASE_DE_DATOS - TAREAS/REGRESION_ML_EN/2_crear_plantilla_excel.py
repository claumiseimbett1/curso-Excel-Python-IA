"""
Script 2: Crear Plantilla Excel para Predicciones
==================================================
Este script crea un archivo Excel con las columnas necesarias para hacer predicciones.

INSTRUCCIONES:
1. Ejecuta este script: python 2_crear_plantilla_excel.py
2. Se creará el archivo: Plantilla_Prediccion_Consumo.xlsx
3. Llena los datos en el Excel
4. Ejecuta el script de predicción (3_predecir_en_excel.py)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
import os

def crear_plantilla_excel():
    """Crea un archivo Excel plantilla para ingresar nuevos datos"""

    print("=" * 60)
    print("CREANDO PLANTILLA EXCEL PARA PREDICCIONES")
    print("=" * 60)

    # Cargar información del modelo
    if not os.path.exists('model_info.json'):
        print("\n❌ ERROR: No se encuentra model_info.json")
        print("\n" + "=" * 60)
        print("DEBES EJECUTAR PRIMERO EL SCRIPT 1")
        print("=" * 60)
        print("1. Abre tu notebook: prediccion_biomasa_ml.ipynb")
        print("2. Ejecuta todas las celdas (Kernel → Restart & Run All)")
        print("3. Agrega una celda al final con el código de guardar modelo")
        print("4. Ejecuta esa celda para generar model_info.json")
        print("5. Luego ejecuta este script nuevamente")
        print("=" * 60)
        return None

    try:
        with open('model_info.json', 'r', encoding='utf-8') as f:
            model_info = json.load(f)
    except json.JSONDecodeError:
        print("\n❌ ERROR: El archivo model_info.json está corrupto")
        print("   Elimínalo y vuelve a ejecutar el Script 1")
        return None

    # Verificar estructura del JSON
    claves_requeridas = ['model_name', 'metricas', 'feature_names', 'fecha_entrenamiento']
    claves_faltantes = [clave for clave in claves_requeridas if clave not in model_info]

    if claves_faltantes:
        print("\n❌ ERROR: El archivo model_info.json no tiene la estructura correcta")
        print(f"   Faltan claves: {claves_faltantes}")
        print("\n   Esto significa que no usaste la función guardar_modelo_entrenado()")
        print("   Por favor, ejecuta el Script 1 correctamente")
        return None

    # Validar que metricas tenga las claves necesarias
    metricas_requeridas = ['R2_test', 'RMSE_test', 'MAE_test']
    if 'metricas' in model_info:
        metricas_faltantes = [m for m in metricas_requeridas if m not in model_info['metricas']]
        if metricas_faltantes:
            print(f"\n⚠ ADVERTENCIA: Faltan métricas: {metricas_faltantes}")
            # Agregar valores por defecto
            for metrica in metricas_faltantes:
                model_info['metricas'][metrica] = 0.0

    feature_names = model_info.get('feature_names', [])

    if not feature_names:
        print("\n❌ ERROR: No hay nombres de variables en model_info.json")
        print("   El modelo no fue guardado correctamente")
        print("   Por favor, ejecuta el Script 1 correctamente")
        return None

    print(f"✓ Modelo cargado: {model_info.get('model_name', 'Desconocido')}")
    print(f"✓ Variables encontradas: {len(feature_names)}")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos para Predicción"

    # Estilos
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    prediction_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    prediction_font = Font(bold=True, color="000000", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws['A1'] = 'PLANTILLA DE PREDICCIÓN DE CONSUMO ENERGÉTICO'
    ws['A1'].font = Font(bold=True, size=14, color="1F4788")
    ws.merge_cells('A1:E1')

    ws['A2'] = f'Modelo: {model_info["model_name"]} | R²: {model_info["metricas"]["R2_test"]:.4f}'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')

    ws['A3'] = f'Creado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = Font(italic=True, size=9)
    ws.merge_cells('A3:E3')

    # Encabezados de columnas (fila 5)
    row = 5
    ws[f'A{row}'] = 'ID'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].border = border

    # Variables predictoras
    for col_idx, feature in enumerate(feature_names, start=2):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = feature
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    # Columna de predicción (amarilla para distinguir)
    pred_col = len(feature_names) + 2
    cell = ws.cell(row=row, column=pred_col)
    cell.value = 'Consumo_kWh_Mensual_Predicho'
    cell.font = prediction_font
    cell.fill = prediction_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

    # Agregar filas de ejemplo (fila 6 en adelante)
    ejemplos = [
        {'ID': 1, 'Sector': 'Residencial', 'Estrato': 3, 'Ciudad': 'Montería', 'Area_m2': 80, 'Puede_Pagar_Solar': 'No'},
        {'ID': 2, 'Sector': 'Comercial', 'Estrato': 5, 'Ciudad': 'Sahagún', 'Area_m2': 150, 'Puede_Pagar_Solar': 'Sí'},
        {'ID': 3, 'Sector': 'Residencial', 'Estrato': 2, 'Ciudad': 'Planeta Rica', 'Area_m2': 100, 'Puede_Pagar_Solar': 'No'},
    ]

    row = 6
    for ejemplo in ejemplos:
        ws[f'A{row}'] = ejemplo['ID']
        ws[f'A{row}'].border = border
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Valores de ejemplo (ajusta según tus variables)
        # Esto asume que feature_names tiene el orden: Sector, Estrato, Ciudad, Area_m2, Puede_Pagar_Solar
        valores_ejemplo = [
            ejemplo.get('Sector', 'Residencial'),
            ejemplo.get('Estrato', 3),
            ejemplo.get('Ciudad', 'Montería'),
            ejemplo.get('Area_m2', 100),
            ejemplo.get('Puede_Pagar_Solar', 'No')
        ]

        for col_idx, valor in enumerate(valores_ejemplo[:len(feature_names)], start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = valor
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Celda de predicción vacía
        cell = ws.cell(row=row, column=pred_col)
        cell.value = ""
        cell.border = border
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

        row += 1

    # Agregar filas vacías para que el usuario pueda agregar datos
    for i in range(10):
        ws[f'A{row}'] = row - 5
        ws[f'A{row}'].border = border
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        for col_idx in range(2, pred_col + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = ""
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if col_idx == pred_col:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        row += 1

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    for col_idx in range(2, pred_col):
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = 20
    ws.column_dimensions[chr(64 + pred_col)].width = 18

    # Instrucciones en otra hoja
    ws_inst = wb.create_sheet("Instrucciones")
    ws_inst['A1'] = 'INSTRUCCIONES DE USO'
    ws_inst['A1'].font = Font(bold=True, size=14, color="1F4788")

    instrucciones = [
        "",
        "1. LLENAR DATOS:",
        "   - En la hoja 'Datos para Predicción', llena las columnas con tus datos",
        "   - NO modifiques los encabezados (fila 5)",
        "   - NO llenes la columna 'Consumo_kWh_Mensual_Predicho' (se llenará automáticamente)",
        "",
        "2. EJECUTAR PREDICCIÓN:",
        "   - Guarda este archivo Excel",
        "   - Ejecuta el script: python 3_predecir_en_excel.py",
        "   - El script leerá este archivo, hará las predicciones, y las escribirá",
        "",
        "3. VER RESULTADOS:",
        "   - Abre nuevamente este archivo Excel",
        "   - La columna 'Consumo_kWh_Mensual_Predicho' tendrá las predicciones",
        "",
        f"INFORMACIÓN DEL MODELO:",
        f"  - Modelo: {model_info['model_name']}",
        f"  - R² Score: {model_info['metricas']['R2_test']:.4f}",
        f"  - RMSE: {model_info['metricas']['RMSE_test']:.2f}",
        f"  - MAE: {model_info['metricas']['MAE_test']:.2f}",
        f"  - Entrenado: {model_info['fecha_entrenamiento']}",
        "",
        "VARIABLES REQUERIDAS:",
    ]

    for var in feature_names:
        instrucciones.append(f"  • {var}")

    row = 1
    for inst in instrucciones:
        ws_inst[f'A{row}'] = inst
        if inst.startswith('INFORMACIÓN') or inst.startswith('VARIABLES'):
            ws_inst[f'A{row}'].font = Font(bold=True, size=11)
        row += 1

    ws_inst.column_dimensions['A'].width = 80

    # Guardar archivo
    filename = 'Plantilla_Prediccion_Consumo.xlsx'
    wb.save(filename)

    print(f"\n✓ Plantilla creada: {filename}")
    print(f"\nVariables requeridas ({len(feature_names)}):")
    for i, var in enumerate(feature_names, 1):
        print(f"  {i}. {var}")

    print("\n" + "=" * 60)
    print("PRÓXIMOS PASOS:")
    print("=" * 60)
    print("1. Abre el archivo: Plantilla_Prediccion_Consumo.xlsx")
    print("2. Llena tus datos en las columnas correspondientes")
    print("3. Guarda el archivo")
    print("4. Ejecuta: python 3_predecir_en_excel.py")
    print("=" * 60)

    return filename


if __name__ == "__main__":
    crear_plantilla_excel()
