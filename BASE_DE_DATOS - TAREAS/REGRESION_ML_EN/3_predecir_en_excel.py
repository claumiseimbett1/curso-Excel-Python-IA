"""
Script 3: Predicción Automática en Excel
=========================================
Este script lee el archivo Excel con nuevos datos, hace predicciones,
y escribe los resultados en el MISMO archivo.

INSTRUCCIONES:
1. Asegúrate de haber llenado el Excel: Plantilla_Prediccion_Consumo.xlsx
2. Ejecuta este script: python 3_predecir_en_excel.py
3. Las predicciones se escribirán en la columna 'Consumo_kWh_Mensual_Predicho'
"""

import pickle
import json
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import sys

def cargar_modelo():
    """Carga el modelo entrenado, scaler e información"""

    print("=" * 70)
    print("SISTEMA DE PREDICCIÓN DE CONSUMO ENERGÉTICO")
    print("=" * 70)

    # Verificar archivos necesarios
    archivos_requeridos = ['best_model.pkl', 'scaler.pkl', 'model_info.json']
    faltantes = [f for f in archivos_requeridos if not os.path.exists(f)]

    if faltantes:
        print("\n❌ ERROR: Faltan archivos necesarios:")
        for archivo in faltantes:
            print(f"   - {archivo}")
        print("\nPor favor, ejecuta primero el script 1_guardar_modelo.py")
        return None, None, None

    # Cargar modelo
    with open('best_model.pkl', 'rb') as f:
        model = pickle.load(f)
    print("✓ Modelo cargado")

    # Cargar scaler
    with open('scaler.pkl', 'rb') as f:
        scaler = pickle.load(f)
    print("✓ Scaler cargado")

    # Cargar información
    with open('model_info.json', 'r', encoding='utf-8') as f:
        info = json.load(f)
    print("✓ Información cargada")

    print(f"\nModelo: {info['model_name']}")
    print(f"R² Score: {info['metricas']['R2_test']:.4f}")
    print(f"RMSE: {info['metricas']['RMSE_test']:.2f}")
    print(f"MAE: {info['metricas']['MAE_test']:.2f}")

    return model, scaler, info


def leer_datos_excel(filename, feature_names):
    """Lee los datos del archivo Excel"""

    print("\n" + "=" * 70)
    print("LEYENDO DATOS DEL EXCEL")
    print("=" * 70)

    if not os.path.exists(filename):
        print(f"\n❌ ERROR: No se encuentra el archivo {filename}")
        print("Por favor, ejecuta primero el script 2_crear_plantilla_excel.py")
        return None

    # Leer Excel (saltando las primeras 4 filas que son título)
    df = pd.read_excel(filename, sheet_name='Datos para Predicción', header=4)

    print(f"✓ Archivo leído: {filename}")
    print(f"  Total de filas: {len(df)}")

    # Filtrar filas vacías (donde todas las columnas predictoras son NaN)
    df_filtrado = df.dropna(how='all', subset=feature_names)

    print(f"  Filas con datos: {len(df_filtrado)}")

    if len(df_filtrado) == 0:
        print("\n❌ ERROR: No hay datos para procesar")
        print("Por favor, llena los datos en el Excel y vuelve a ejecutar")
        return None

    # Verificar columnas requeridas
    columnas_faltantes = [col for col in feature_names if col not in df.columns]
    if columnas_faltantes:
        print(f"\n❌ ERROR: Faltan columnas en el Excel:")
        for col in columnas_faltantes:
            print(f"   - {col}")
        return None

    print(f"\n✓ Columnas verificadas: {len(feature_names)} variables")

    return df_filtrado


def preprocesar_datos(df, feature_names, scaler):
    """Preprocesa los datos para predicción"""

    print("\n" + "=" * 70)
    print("PREPROCESANDO DATOS")
    print("=" * 70)

    # Seleccionar solo las columnas necesarias
    X = df[feature_names].copy()

    # Verificar valores faltantes
    valores_faltantes = X.isnull().sum()
    if valores_faltantes.any():
        print("\n⚠ ADVERTENCIA: Hay valores faltantes:")
        for col, count in valores_faltantes[valores_faltantes > 0].items():
            print(f"   - {col}: {count} valores faltantes")
        print("\n  Se rellenarán con la mediana de cada columna")

        # Rellenar con mediana
        for col in X.columns:
            if X[col].isnull().any():
                if X[col].dtype in ['float64', 'int64']:
                    X[col].fillna(X[col].median(), inplace=True)
                else:
                    X[col].fillna(X[col].mode()[0] if not X[col].mode().empty else 'Franco', inplace=True)

    # Codificar variables categóricas (si existen)
    categorical_cols = X.select_dtypes(include=['object']).columns
    if len(categorical_cols) > 0:
        print(f"\n✓ Codificando variables categóricas: {list(categorical_cols)}")

        # Mapeos para las variables categóricas
        sector_map = {'Residencial': 0, 'Comercial': 1, 'Industrial': 2}
        ciudad_map = {'Montería': 0, 'Sahagún': 1, 'Planeta Rica': 2, 'Cereté': 3, 'Lorica': 4}
        puede_pagar_map = {'No': 0, 'Sí': 1, 'Si': 1}

        for col in categorical_cols:
            if col == 'Sector':
                X[col] = X[col].map(sector_map)
                X[col].fillna(0, inplace=True)  # Residencial por defecto
            elif col == 'Ciudad':
                X[col] = X[col].map(ciudad_map)
                X[col].fillna(0, inplace=True)  # Montería por defecto
            elif col == 'Puede_Pagar_Solar':
                X[col] = X[col].map(puede_pagar_map)
                X[col].fillna(0, inplace=True)  # No por defecto
            else:
                # Para otras categóricas, usar label encoding simple
                X[col] = pd.Categorical(X[col]).codes

    print(f"✓ Datos preprocesados: {X.shape}")

    # Escalar datos
    X_scaled = scaler.transform(X)
    print("✓ Datos escalados")

    return X_scaled


def hacer_predicciones(model, X_scaled):
    """Hace las predicciones usando el modelo"""

    print("\n" + "=" * 70)
    print("HACIENDO PREDICCIONES")
    print("=" * 70)

    predicciones = model.predict(X_scaled)

    print(f"✓ Predicciones realizadas: {len(predicciones)} valores")
    print(f"\n  Estadísticas de predicciones:")
    print(f"    - Mínimo: {predicciones.min():.2f}")
    print(f"    - Máximo: {predicciones.max():.2f}")
    print(f"    - Promedio: {predicciones.mean():.2f}")
    print(f"    - Mediana: {np.median(predicciones):.2f}")

    return predicciones


def escribir_resultados(filename, predicciones, df_original):
    """Escribe las predicciones en el mismo archivo Excel"""

    print("\n" + "=" * 70)
    print("ESCRIBIENDO RESULTADOS EN EXCEL")
    print("=" * 70)

    # Cargar el workbook existente
    wb = load_workbook(filename)
    ws = wb['Datos para Predicción']

    # Encontrar la columna de predicciones
    header_row = 5
    pred_col = None

    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value == 'Consumo_kWh_Mensual_Predicho':
            pred_col = col
            break

    if pred_col is None:
        print("❌ ERROR: No se encuentra la columna 'Consumo_kWh_Mensual_Predicho'")
        return False

    # Escribir predicciones (empezando en fila 6)
    data_start_row = 6
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for i, pred in enumerate(predicciones):
        row = data_start_row + df_original.index[i]
        cell = ws.cell(row=row, column=pred_col)
        cell.value = round(float(pred), 2)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = fill_verde

    # Agregar marca de tiempo
    ws['A3'] = f'Última predicción: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = Font(italic=True, size=9, color="006400")

    # Guardar archivo
    wb.save(filename)

    print(f"✓ Resultados escritos en: {filename}")
    print(f"  Columna: Consumo_kWh_Mensual_Predicho")
    print(f"  Filas actualizadas: {len(predicciones)}")

    return True


def main():
    """Función principal"""

    filename = 'Plantilla_Prediccion_Consumo.xlsx'

    # 1. Cargar modelo
    model, scaler, info = cargar_modelo()
    if model is None:
        return

    feature_names = info['feature_names']

    # 2. Leer datos
    df = leer_datos_excel(filename, feature_names)
    if df is None:
        return

    # 3. Preprocesar
    try:
        X_scaled = preprocesar_datos(df, feature_names, scaler)
    except Exception as e:
        print(f"\n❌ ERROR al preprocesar datos: {str(e)}")
        return

    # 4. Predecir
    try:
        predicciones = hacer_predicciones(model, X_scaled)
    except Exception as e:
        print(f"\n❌ ERROR al hacer predicciones: {str(e)}")
        return

    # 5. Escribir resultados
    try:
        exito = escribir_resultados(filename, predicciones, df)
    except Exception as e:
        print(f"\n❌ ERROR al escribir resultados: {str(e)}")
        return

    if exito:
        print("\n" + "=" * 70)
        print("✓ ¡PROCESO COMPLETADO EXITOSAMENTE!")
        print("=" * 70)
        print(f"\nAbre el archivo {filename} para ver las predicciones")
        print("Las predicciones están en la columna 'Consumo_kWh_Mensual_Predicho' (fondo verde)")
        print("\n" + "=" * 70)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO: {str(e)}")
        import traceback
        traceback.print_exc()
