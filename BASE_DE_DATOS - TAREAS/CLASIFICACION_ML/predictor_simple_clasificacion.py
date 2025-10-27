"""
CLASIFICADOR SIMPLIFICADO PARA EXCEL
=====================================
Script simplificado para clasificar biomasa desde Excel sin complicaciones.

Uso: python predictor_simple_clasificacion.py
"""

import pickle
import json
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os


def clasificar_biomasa_simple(archivo_excel='Plantilla_Clasificacion_Biomasa.xlsx'):
    """Función principal para clasificar biomasa"""

    print("\n" + "🔍" * 35)
    print("     CLASIFICADOR DE BIOMASA - VERSIÓN SIMPLE")
    print("🔍" * 35 + "\n")

    # Verificar archivos
    archivos = {
        'modelo': 'best_model_clasificacion.pkl',
        'scaler': 'scaler_clasificacion.pkl',
        'encoder': 'label_encoder_clasificacion.pkl',
        'info': 'model_info_clasificacion.json',
        'excel': archivo_excel
    }

    for nombre, archivo in archivos.items():
        if not os.path.exists(archivo):
            print(f"   ❌ No se encuentra: {archivo}")
            if nombre != 'excel':
                print(f"\n💡 Ejecuta desde el notebook: %run 1_guardar_modelo_clasificacion.py")
            else:
                print(f"\n💡 Ejecuta: python 2_crear_plantilla_excel_clasificacion.py")
            return

    print("📋 Cargando modelo...")

    # Cargar modelo
    with open('best_model_clasificacion.pkl', 'rb') as f:
        modelo = pickle.load(f)

    with open('scaler_clasificacion.pkl', 'rb') as f:
        scaler = pickle.load(f)

    with open('label_encoder_clasificacion.pkl', 'rb') as f:
        le_target = pickle.load(f)

    with open('model_info_clasificacion.json', 'r') as f:
        info = json.load(f)

    features = info['feature_names']
    classes = info['classes']

    print(f"   ✓ Modelo: {info['model_name']}")
    print(f"   ✓ Clases: {', '.join(classes)}")
    print(f"   ✓ Accuracy: {info['metricas']['Accuracy_test']:.4f}")

    # Leer Excel
    print(f"\n📊 Leyendo Excel: {archivo_excel}...")
    df = pd.read_excel(archivo_excel, sheet_name='Datos para Clasificación', header=4)

    filas_con_datos = df[features].notna().all(axis=1)
    df_valido = df[filas_con_datos].copy()

    if len(df_valido) == 0:
        print("\n   ⚠️  NO HAY DATOS PARA CLASIFICAR")
        return

    print(f"   ✓ Filas con datos: {len(df_valido)}")

    # Preprocesar
    print("\n🔧 Preprocesando...")
    X = df_valido[features].copy()

    if 'Tipo_suelo' in X.columns:
        tipo_suelo_map = {'Arenoso': 0, 'Arcilloso': 1, 'Franco': 2}
        X['Tipo_suelo'] = X['Tipo_suelo'].map(tipo_suelo_map)
        X['Tipo_suelo'].fillna(2, inplace=True)

    X_scaled = scaler.transform(X)

    # Clasificar
    print("\n🎯 Clasificando...")
    predicciones_encoded = modelo.predict(X_scaled)
    predicciones = le_target.inverse_transform(predicciones_encoded)

    unique, counts = np.unique(predicciones, return_counts=True)
    print(f"\n   📊 Resultados:")
    for clase, count in zip(unique, counts):
        print(f"      • {clase}: {count} muestras ({count/len(predicciones)*100:.1f}%)")

    # Escribir en Excel
    print(f"\n💾 Escribiendo en Excel...")

    wb = load_workbook(archivo_excel)
    ws = wb['Datos para Clasificación']

    # Encontrar columna
    header_row = 5
    pred_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == 'Categoria_Predicha':
            pred_col = col
            break

    # Colores
    colores = {
        'Baja': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'Media': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'Alta': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    }

    # Escribir
    data_start_row = 6
    border = Border(
        left=Side(style='thin', color='999999'),
        right=Side(style='thin', color='999999'),
        top=Side(style='thin', color='999999'),
        bottom=Side(style='thin', color='999999')
    )

    for idx, pred in zip(df_valido.index, predicciones):
        row = data_start_row + idx
        cell = ws.cell(row=row, column=pred_col)
        cell.value = pred
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = colores.get(pred, PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"))
        cell.border = border
        cell.font = Font(bold=True, size=11)

    # Timestamp
    ws['A3'] = f'✓ Última clasificación: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = Font(italic=True, size=10, color="006400", bold=True)

    wb.save(archivo_excel)

    print(f"   ✓ Archivo guardado")

    print("\n" + "=" * 70)
    print("✅ ¡CLASIFICACIÓN COMPLETADA!")
    print("=" * 70)
    print(f"\n📂 Abre: {archivo_excel}")
    print("🎯 Las clasificaciones están en la columna 'Categoria_Predicha'")
    print("\n🎨 Colores:")
    print("   🟢 Verde  = Biomasa Alta")
    print("   🟡 Amarillo = Biomasa Media")
    print("   🔴 Rojo   = Biomasa Baja")
    print("\n" + "=" * 70 + "\n")


def clasificar_valores_directos(**kwargs):
    """Clasificar biomasa con valores específicos desde Python"""

    print("\n🎯 CLASIFICACIÓN DIRECTA DE BIOMASA")
    print("=" * 50)

    # Cargar modelo
    if not os.path.exists('best_model_clasificacion.pkl'):
        print("❌ Error: Ejecuta primero el script 1")
        return None

    with open('best_model_clasificacion.pkl', 'rb') as f:
        modelo = pickle.load(f)

    with open('scaler_clasificacion.pkl', 'rb') as f:
        scaler = pickle.load(f)

    with open('label_encoder_clasificacion.pkl', 'rb') as f:
        le_target = pickle.load(f)

    with open('model_info_clasificacion.json', 'r') as f:
        info = json.load(f)

    features = info['feature_names']

    # Verificar valores
    faltantes = [f for f in features if f not in kwargs]
    if faltantes:
        print(f"❌ Faltan valores para: {', '.join(faltantes)}")
        return None

    # Crear DataFrame
    data = {f: [kwargs[f]] for f in features}
    df = pd.DataFrame(data)

    # Codificar
    if 'Tipo_suelo' in df.columns:
        tipo_suelo_map = {'Arenoso': 0, 'Arcilloso': 1, 'Franco': 2}
        df['Tipo_suelo'] = df['Tipo_suelo'].map(tipo_suelo_map)

    # Clasificar
    X_scaled = scaler.transform(df)
    prediccion_encoded = modelo.predict(X_scaled)[0]
    prediccion = le_target.inverse_transform([prediccion_encoded])[0]

    print("\n📊 Valores de entrada:")
    for feature, valor in kwargs.items():
        print(f"   • {feature}: {valor}")

    print(f"\n🎯 CLASIFICACIÓN: {prediccion}")
    print("=" * 50 + "\n")

    return prediccion


if __name__ == "__main__":
    try:
        clasificar_biomasa_simple()
    except KeyboardInterrupt:
        print("\n\n⚠️  Proceso cancelado")
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
