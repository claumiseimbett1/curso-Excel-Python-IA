"""
Script 3: Clasificación Automática en Excel
===========================================
Este script lee el archivo Excel con nuevos datos, hace clasificaciones,
y escribe los resultados en el MISMO archivo.

INSTRUCCIONES:
1. Asegúrate de haber llenado el Excel: Plantilla_Clasificacion_Biomasa.xlsx
2. Ejecuta este script: python 3_predecir_en_excel_clasificacion.py
3. Las clasificaciones se escribirán en la columna 'Categoria_Predicha'
"""

import pickle
import json
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import sys


def cargar_modelo():
    """Carga el modelo, scaler y label encoder"""

    print("=" * 70)
    print("SISTEMA DE CLASIFICACIÓN DE BIOMASA")
    print("=" * 70)

    # Verificar archivos necesarios
    archivos_requeridos = ['best_model_clasificacion.pkl', 'scaler_clasificacion.pkl',
                           'label_encoder_clasificacion.pkl', 'model_info_clasificacion.json']
    faltantes = [f for f in archivos_requeridos if not os.path.exists(f)]

    if faltantes:
        print("\n❌ ERROR: Faltan archivos necesarios:")
        for archivo in faltantes:
            print(f"   - {archivo}")
        print("\nPor favor, ejecuta primero desde el notebook:")
        print("  %run 1_guardar_modelo_clasificacion.py")
        return None, None, None, None

    # Cargar modelo
    with open('best_model_clasificacion.pkl', 'rb') as f:
        model = pickle.load(f)
    print("✓ Modelo cargado")

    # Cargar scaler
    with open('scaler_clasificacion.pkl', 'rb') as f:
        scaler = pickle.load(f)
    print("✓ Scaler cargado")

    # Cargar label encoder
    with open('label_encoder_clasificacion.pkl', 'rb') as f:
        le_target = pickle.load(f)
    print("✓ LabelEncoder cargado")

    # Cargar información
    with open('model_info_clasificacion.json', 'r', encoding='utf-8') as f:
        info = json.load(f)
    print("✓ Información cargada")

    print(f"\nModelo: {info['model_name']}")
    print(f"Clases: {', '.join(info['classes'])}")
    print(f"Accuracy: {info['metricas']['Accuracy_test']:.4f}")
    print(f"F1-Score: {info['metricas']['F1_test']:.4f}")

    return model, scaler, le_target, info


def leer_datos_excel(filename, feature_names):
    """Lee los datos del archivo Excel"""

    print("\n" + "=" * 70)
    print("LEYENDO DATOS DEL EXCEL")
    print("=" * 70)

    if not os.path.exists(filename):
        print(f"\n❌ ERROR: No se encuentra el archivo {filename}")
        print("Por favor, ejecuta primero: python 2_crear_plantilla_excel_clasificacion.py")
        return None

    # Leer Excel (header en fila 5, datos empiezan en fila 6)
    df = pd.read_excel(filename, sheet_name='Datos para Clasificación', header=4)

    print(f"✓ Archivo leído: {filename}")
    print(f"  Total de filas: {len(df)}")

    # Filtrar filas vacías
    df_filtrado = df.dropna(how='all', subset=feature_names)

    print(f"  Filas con datos: {len(df_filtrado)}")

    if len(df_filtrado) == 0:
        print("\n❌ ERROR: No hay datos para clasificar")
        print("Por favor, llena los datos en el Excel y vuelve a ejecutar")
        return None

    # Verificar columnas requeridas
    columnas_faltantes = [col for col in feature_names if col not in df.columns]
    if columnas_faltantes:
        print(f"\n❌ ERROR: Faltan columnas en el Excel:")
        for col in columnas_faltantes:
            print(f"   - {col}")
        return None

    print(f"\n✓ Columnas verificadas: {len(feature_names)} variables")

    return df_filtrado


def preprocesar_datos(df, feature_names, scaler):
    """Preprocesa los datos para clasificación"""

    print("\n" + "=" * 70)
    print("PREPROCESANDO DATOS")
    print("=" * 70)

    # Seleccionar solo las columnas necesarias
    X = df[feature_names].copy()

    # Verificar valores faltantes
    valores_faltantes = X.isnull().sum()
    if valores_faltantes.any():
        print("\n⚠ ADVERTENCIA: Hay valores faltantes:")
        for col, count in valores_faltantes[valores_faltantes > 0].items():
            print(f"   - {col}: {count} valores faltantes")
        print("\n  Se rellenarán con la mediana de cada columna")

        # Rellenar con mediana
        for col in X.columns:
            if X[col].isnull().any():
                if X[col].dtype in ['float64', 'int64']:
                    X[col].fillna(X[col].median(), inplace=True)
                else:
                    X[col].fillna(X[col].mode()[0] if not X[col].mode().empty else 'Franco', inplace=True)

    # Codificar variables categóricas
    categorical_cols = X.select_dtypes(include=['object']).columns
    if len(categorical_cols) > 0:
        print(f"\n✓ Codificando variables categóricas: {list(categorical_cols)}")

        # Mapeo para Tipo_suelo
        tipo_suelo_map = {'Arenoso': 0, 'Arcilloso': 1, 'Franco': 2}

        for col in categorical_cols:
            if col == 'Tipo_suelo':
                X[col] = X[col].map(tipo_suelo_map)
                X[col].fillna(2, inplace=True)
            else:
                X[col] = pd.Categorical(X[col]).codes

    print(f"✓ Datos preprocesados: {X.shape}")

    # Escalar datos
    X_scaled = scaler.transform(X)
    print("✓ Datos escalados")

    return X_scaled


def hacer_clasificacion(model, X_scaled, le_target):
    """Hace las clasificaciones usando el modelo"""

    print("\n" + "=" * 70)
    print("HACIENDO CLASIFICACIONES")
    print("=" * 70)

    # Predicciones
    predicciones_encoded = model.predict(X_scaled)

    # Decodificar predicciones
    predicciones = le_target.inverse_transform(predicciones_encoded)

    print(f"✓ Clasificaciones realizadas: {len(predicciones)} valores")

    # Estadísticas
    unique, counts = np.unique(predicciones, return_counts=True)
    print(f"\n  Distribución de clasificaciones:")
    for clase, count in zip(unique, counts):
        print(f"    - {clase}: {count} ({count/len(predicciones)*100:.1f}%)")

    # Probabilidades si están disponibles
    if hasattr(model, "predict_proba"):
        try:
            # Para GridSearchCV, usar best_estimator_
            if hasattr(model, 'best_estimator_'):
                probas = model.best_estimator_.predict_proba(X_scaled)
            else:
                probas = model.predict_proba(X_scaled)

            # Confianza promedio
            max_probas = probas.max(axis=1)
            print(f"\n  Confianza de las predicciones:")
            print(f"    - Promedio: {max_probas.mean()*100:.1f}%")
            print(f"    - Mínimo: {max_probas.min()*100:.1f}%")
            print(f"    - Máximo: {max_probas.max()*100:.1f}%")

            return predicciones, max_probas
        except:
            return predicciones, None
    else:
        return predicciones, None


def escribir_resultados(filename, predicciones, probas, df_original):
    """Escribe las clasificaciones en el mismo archivo Excel"""

    print("\n" + "=" * 70)
    print("ESCRIBIENDO RESULTADOS EN EXCEL")
    print("=" * 70)

    # Cargar el workbook existente
    wb = load_workbook(filename)
    ws = wb['Datos para Clasificación']

    # Encontrar la columna de predicciones
    header_row = 5
    pred_col = None

    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value == 'Categoria_Predicha':
            pred_col = col
            break

    if pred_col is None:
        print("❌ ERROR: No se encuentra la columna 'Categoria_Predicha'")
        return False

    # Colores por categoría
    colores = {
        'Baja': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'Media': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'Alta': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    }

    # Escribir predicciones
    data_start_row = 6

    for i, pred in enumerate(predicciones):
        row = data_start_row + df_original.index[i]
        cell = ws.cell(row=row, column=pred_col)
        cell.value = pred
        cell.alignment = Alignment(horizontal='center')

        # Color según categoría
        if pred in colores:
            cell.fill = colores[pred]
        else:
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        cell.font = Font(bold=True, size=11)

    # Agregar marca de tiempo
    ws['A3'] = f'✓ Última clasificación: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = Font(italic=True, size=9, color="006400", bold=True)

    # Guardar archivo
    wb.save(filename)

    print(f"✓ Resultados escritos en: {filename}")
    print(f"  Columna: Categoria_Predicha")
    print(f"  Filas actualizadas: {len(predicciones)}")

    # Leyenda de colores
    print(f"\n  Colores:")
    print(f"    🟢 Verde: Biomasa Alta")
    print(f"    🟡 Amarillo: Biomasa Media")
    print(f"    🔴 Rojo: Biomasa Baja")

    return True


def main():
    """Función principal"""

    filename = 'Plantilla_Clasificacion_Biomasa.xlsx'

    # 1. Cargar modelo
    model, scaler, le_target, info = cargar_modelo()
    if model is None:
        return

    feature_names = info['feature_names']

    # 2. Leer datos
    df = leer_datos_excel(filename, feature_names)
    if df is None:
        return

    # 3. Preprocesar
    try:
        X_scaled = preprocesar_datos(df, feature_names, scaler)
    except Exception as e:
        print(f"\n❌ ERROR al preprocesar datos: {str(e)}")
        return

    # 4. Clasificar
    try:
        predicciones, probas = hacer_clasificacion(model, X_scaled, le_target)
    except Exception as e:
        print(f"\n❌ ERROR al hacer clasificaciones: {str(e)}")
        return

    # 5. Escribir resultados
    try:
        exito = escribir_resultados(filename, predicciones, probas, df)
    except Exception as e:
        print(f"\n❌ ERROR al escribir resultados: {str(e)}")
        return

    if exito:
        print("\n" + "=" * 70)
        print("✅ ¡CLASIFICACIÓN COMPLETADA EXITOSAMENTE!")
        print("=" * 70)
        print(f"\nAbre el archivo {filename} para ver las clasificaciones")
        print("Las clasificaciones están en la columna 'Categoria_Predicha' (con colores)")
        print("\n" + "=" * 70)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO: {str(e)}")
        import traceback
        traceback.print_exc()
