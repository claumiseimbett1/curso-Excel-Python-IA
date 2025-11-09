"""
Script 2: Crear Plantilla Excel para Clasificación de Biomasa
==============================================================
Este script crea un archivo Excel con las columnas necesarias para clasificar biomasa.

INSTRUCCIONES:
1. Ejecuta este script: python 2_crear_plantilla_excel_clasificacion.py
2. Se creará el archivo: Plantilla_Clasificacion_Biomasa.xlsx
3. Llena los datos en el Excel
4. Ejecuta el script de clasificación (3_predecir_en_excel_clasificacion.py)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
import os


def crear_plantilla_excel_clasificacion():
    """Crea un archivo Excel plantilla para clasificación de biomasa"""

    print("\n" + "=" * 70)
    print("CREANDO PLANTILLA EXCEL PARA CLASIFICACIÓN DE BIOMASA")
    print("=" * 70 + "\n")

    # Cargar información del modelo
    if not os.path.exists('model_info_clasificacion.json'):
        print("❌ ERROR: No se encuentra model_info_clasificacion.json")
        print("\n💡 SOLUCIÓN:")
        print("1. Abre el notebook: 0_clasificacion_biomasa_ml.ipynb")
        print("2. Ejecuta todas las celdas (Kernel → Restart & Run All)")
        print("3. Agrega una celda al final con:")
        print("   %run 1_guardar_modelo_clasificacion.py")
        print("4. Ejecuta esa celda")
        print("5. Luego ejecuta este script nuevamente")
        return None

    with open('model_info_clasificacion.json', 'r', encoding='utf-8') as f:
        model_info = json.load(f)

    feature_names = model_info['variables_predictoras']
    classes = model_info['clases']

    print(f"✓ Modelo cargado: {model_info['modelo']}")
    print(f"✓ Clases: {', '.join(classes)}")
    print(f"✓ Variables: {len(feature_names)}\n")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos para Clasificación"

    # Estilos
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    prediction_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    prediction_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws['A1'] = '🔍 CLASIFICADOR DE BIOMASA - MACHINE LEARNING'
    ws['A1'].font = Font(bold=True, size=14, color="1565C0")
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Información del modelo
    ws['A2'] = f'📊 Modelo: {model_info["modelo"]} | Accuracy: {model_info["metricas"]["accuracy_test"]:.4f} | F1-Score: {model_info["metricas"]["f1_test"]:.4f}'
    ws['A2'].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells('A2:F2')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Clases disponibles
    ws['A3'] = f'🎯 Clases: {", ".join(classes)}'
    ws['A3'].font = Font(bold=True, size=10, color="FF6F00")
    ws.merge_cells('A3:F3')
    ws['A3'].alignment = Alignment(horizontal='center')

    ws['A4'] = ''  # Fila vacía

    # Encabezados de columnas (fila 5)
    header_row = 5

    # ID
    ws[f'A{header_row}'] = 'ID'
    ws[f'A{header_row}'].font = header_font
    ws[f'A{header_row}'].fill = header_fill
    ws[f'A{header_row}'].alignment = Alignment(horizontal='center')
    ws[f'A{header_row}'].border = border

    # Variables predictoras
    for col_idx, feature in enumerate(feature_names, start=2):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = feature
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    # Columna de clasificación
    pred_col = len(feature_names) + 2
    cell = ws.cell(row=header_row, column=pred_col)
    cell.value = 'Categoria_Predicha'
    cell.font = prediction_font
    cell.fill = prediction_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

    ws.row_dimensions[header_row].height = 35

    # Agregar filas vacías para datos (20 filas)
    data_start_row = 6
    for i in range(20):
        row = data_start_row + i

        # ID
        ws[f'A{row}'] = i + 1
        ws[f'A{row}'].border = border
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Columnas de variables
        for col_idx in range(2, pred_col):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = ""
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Columna de predicción
        cell = ws.cell(row=row, column=pred_col)
        cell.value = ""
        cell.border = border
        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 6
    for col_idx in range(2, pred_col):
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = 20
    ws.column_dimensions[chr(64 + pred_col)].width = 18

    # Hoja de Instrucciones
    ws_inst = wb.create_sheet("📖 Instrucciones")

    ws_inst['A1'] = '📖 GUÍA DE USO - CLASIFICADOR DE BIOMASA'
    ws_inst['A1'].font = Font(bold=True, size=14, color="1976D2")
    ws_inst.merge_cells('A1:B1')
    ws_inst['A1'].alignment = Alignment(horizontal='center')
    ws_inst.row_dimensions[1].height = 25

    instrucciones = [
        "",
        "🚀 INSTRUCCIONES DE USO",
        "═" * 80,
        "",
        "1️⃣ LLENAR DATOS",
        "   • Ve a la hoja 'Datos para Clasificación'",
        "   • Ingresa tus datos en las columnas correspondientes",
        "   • NO modifiques los encabezados (fila 5)",
        "   • NO llenes la columna 'Categoria_Predicha' (se llenará automáticamente)",
        "",
        "2️⃣ EJECUTAR CLASIFICACIÓN",
        "   • Guarda este archivo Excel (Ctrl+S)",
        "   • Cierra Excel",
        "   • Abre una terminal o CMD",
        "   • Ejecuta: python 3_predecir_en_excel_clasificacion.py",
        "",
        "3️⃣ VER RESULTADOS",
        "   • Abre nuevamente este Excel",
        "   • La columna 'Categoria_Predicha' tendrá las clasificaciones",
        "   • Valores posibles: " + ", ".join(classes),
        "",
        "",
        "📊 INFORMACIÓN DEL MODELO",
        "═" * 80,
        "",
        f"   Modelo:            {model_info['modelo']}",
        f"   Tipo:              Clasificación",
        f"   Clases:            {', '.join(classes)}",
        f"   Accuracy:          {model_info['metricas']['accuracy_test']:.4f}",
        f"   Precision:         {model_info['metricas']['precision_test']:.4f}",
        f"   Recall:            {model_info['metricas']['recall_test']:.4f}",
        f"   F1-Score:          {model_info['metricas']['f1_test']:.4f}",
        f"   Entrenado:         {model_info['fecha_entrenamiento']}",
        "",
        "",
        "📋 VARIABLES REQUERIDAS",
        "═" * 80,
        "",
    ]

    for var in feature_names:
        instrucciones.append(f"   ✓ {var}")

    instrucciones.extend([
        "",
        "",
        "🎯 INTERPRETACIÓN DE RESULTADOS",
        "═" * 80,
        "",
        "El modelo clasificará cada muestra en una de estas categorías:",
        "",
    ])

    for clase in classes:
        instrucciones.append(f"   • {clase}")

    instrucciones.extend([
        "",
        "La clasificación se basa en los valores de las variables predictoras",
        "que ingreses en el Excel.",
        "",
        "",
        "❓ SOLUCIÓN DE PROBLEMAS",
        "═" * 80,
        "",
        "❌ Error: 'No se encuentra best_model_clasificacion.pkl'",
        "   → Ejecuta: %run 1_guardar_modelo_clasificacion.py (desde el notebook)",
        "",
        "❌ Error: 'No hay datos para clasificar'",
        "   → Llena al menos una fila completa en el Excel",
        "",
        "❌ Las clasificaciones no aparecen",
        "   → Verifica que guardaste el Excel antes de ejecutar el script",
        "   → Cierra Excel antes de ejecutar",
        "",
    ])

    row = 1
    for inst in instrucciones:
        ws_inst[f'A{row}'] = inst

        # Estilos
        if inst.startswith('═'):
            ws_inst[f'A{row}'].font = Font(color="CCCCCC", size=10)
        elif any(inst.startswith(emoji) for emoji in ['🚀', '📊', '📋', '🎯', '❓']):
            ws_inst[f'A{row}'].font = Font(bold=True, size=13, color="1976D2")
            ws_inst[f'A{row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        elif any(inst.startswith(emoji) for emoji in ['1️⃣', '2️⃣', '3️⃣']):
            ws_inst[f'A{row}'].font = Font(bold=True, size=11, color="F57C00")
        elif inst.startswith('   ✓'):
            ws_inst[f'A{row}'].font = Font(size=10, color="2E7D32")
        elif inst.startswith('   •'):
            ws_inst[f'A{row}'].font = Font(size=10, color="555555")
        elif inst.startswith('   →'):
            ws_inst[f'A{row}'].font = Font(italic=True, size=10, color="1976D2")
        else:
            ws_inst[f'A{row}'].font = Font(size=10)

        row += 1

    ws_inst.column_dimensions['A'].width = 100

    # Guardar archivo
    filename = 'Plantilla_Clasificacion_Biomasa.xlsx'
    wb.save(filename)

    print("=" * 70)
    print("✅ PLANTILLA EXCEL CREADA")
    print("=" * 70)
    print(f"\n📂 Archivo creado: {filename}")
    print(f"🎯 Clases posibles: {', '.join(classes)}")
    print(f"📋 Variables requeridas: {len(feature_names)}")

    print(f"\n📋 Lista de variables:")
    for i, var in enumerate(feature_names, 1):
        print(f"   {i}. {var}")

    print("\n" + "=" * 70)
    print("PRÓXIMOS PASOS:")
    print("=" * 70)
    print(f"1. Abre: {filename}")
    print("2. Llena tus datos en las columnas correspondientes")
    print("3. Guarda y cierra el Excel")
    print("4. Ejecuta: python 3_predecir_en_excel_clasificacion.py")
    print("=" * 70 + "\n")

    return filename


if __name__ == "__main__":
    try:
        crear_plantilla_excel_clasificacion()
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
