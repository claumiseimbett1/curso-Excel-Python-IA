"""
Script 4: Crear Excel con Botón de Predicción Integrado
=========================================================
Este script crea un archivo Excel con macros (.xlsm) que incluye un botón
para ejecutar las predicciones directamente desde Excel.

CARACTERÍSTICAS:
- Botón "Predecir Biomasa" integrado en el Excel
- Ejecuta las predicciones con un solo clic
- No necesitas abrir Python ni terminal
- Actualización automática en el mismo Excel

INSTRUCCIONES:
1. Ejecuta este script: python 4_crear_excel_con_boton.py
2. Se creará: Plantilla_Prediccion_Con_Boton.xlsm
3. Abre el archivo Excel
4. Llena los datos
5. Haz clic en el botón "🎯 PREDECIR BIOMASA"
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def crear_codigo_vba():
    """Genera el código VBA que ejecutará las predicciones"""

    vba_code = '''
Sub PredecirBiomasa()
    '=============================================================================
    ' MACRO: Predicción de Biomasa usando Machine Learning
    '=============================================================================
    ' Esta macro ejecuta el script Python que hace las predicciones y actualiza
    ' automáticamente este archivo Excel con los resultados
    '=============================================================================

    Dim pythonPath As String
    Dim scriptPath As String
    Dim command As String
    Dim shell As Object
    Dim waitOnReturn As Boolean: waitOnReturn = True
    Dim windowStyle As Integer: windowStyle = 1  ' 1 = Normal, 0 = Hidden
    Dim resultado As Long

    ' Deshabilitar actualización de pantalla para mejor rendimiento
    Application.ScreenUpdating = False
    Application.DisplayAlerts = False

    ' Mensaje inicial
    MsgBox "Iniciando predicción de biomasa..." & vbCrLf & vbCrLf & _
           "Este proceso puede tomar unos segundos.", vbInformation, "Predictor de Biomasa"

    ' Guardar el archivo antes de ejecutar Python
    ThisWorkbook.Save

    ' Detectar la ruta del archivo actual
    scriptPath = ThisWorkbook.Path

    ' Intentar encontrar Python (prueba varios paths comunes)
    pythonPath = ""

    ' Opción 1: Python en PATH del sistema
    pythonPath = "python"

    ' Construir el comando completo
    ' El script predictor_excel_simple.py debe estar en la misma carpeta
    command = "cmd /c cd /d """ & scriptPath & """ && " & pythonPath & " predictor_excel_simple.py"

    ' Ejecutar el comando
    Set shell = CreateObject("WScript.Shell")

    On Error GoTo ErrorHandler
    resultado = shell.Run(command, windowStyle, waitOnReturn)

    ' Esperar un momento para que Python termine de escribir
    Application.Wait (Now + TimeValue("0:00:02"))

    ' Cerrar el archivo actual (sin guardar porque Python ya lo modificó)
    ThisWorkbook.Close SaveChanges:=False

    ' Reabrir el archivo para mostrar las predicciones
    Dim reopenPath As String
    reopenPath = scriptPath & "\\" & ThisWorkbook.Name
    Workbooks.Open reopenPath

    ' Mensaje de éxito
    MsgBox "✅ ¡Predicciones completadas!" & vbCrLf & vbCrLf & _
           "Las predicciones están en la columna 'Biomasa_Predicha'", _
           vbInformation, "Éxito"

    Application.ScreenUpdating = True
    Application.DisplayAlerts = True

    Exit Sub

ErrorHandler:
    Application.ScreenUpdating = True
    Application.DisplayAlerts = True

    MsgBox "❌ Error al ejecutar la predicción." & vbCrLf & vbCrLf & _
           "Detalles del error:" & vbCrLf & _
           Err.Description & vbCrLf & vbCrLf & _
           "Verifica que:" & vbCrLf & _
           "1. Python está instalado" & vbCrLf & _
           "2. El archivo predictor_excel_simple.py está en la misma carpeta" & vbCrLf & _
           "3. Los archivos del modelo (best_model.pkl, scaler.pkl) existen", _
           vbCritical, "Error"
End Sub


Sub VerificarPython()
    '=============================================================================
    ' MACRO: Verificar instalación de Python
    '=============================================================================

    Dim shell As Object
    Dim resultado As Long

    Set shell = CreateObject("WScript.Shell")
    resultado = shell.Run("cmd /c python --version", 1, True)

    If resultado = 0 Then
        MsgBox "✅ Python está instalado correctamente", vbInformation
    Else
        MsgBox "❌ Python no encontrado" & vbCrLf & vbCrLf & _
               "Por favor, instala Python desde:" & vbCrLf & _
               "https://www.python.org/downloads/", vbCritical
    End If
End Sub
'''

    return vba_code


def crear_excel_con_boton():
    """Crea el archivo Excel con la estructura y prepara para macros"""

    print("\n" + "🚀" * 35)
    print("   CREADOR DE EXCEL CON BOTÓN DE PREDICCIÓN")
    print("🚀" * 35 + "\n")

    # Verificar modelo
    if not os.path.exists('model_info.json'):
        print("❌ Error: No se encuentra model_info.json")
        print("💡 Ejecuta primero: python 1_guardar_modelo.py\n")
        return None

    with open('model_info.json', 'r', encoding='utf-8') as f:
        model_info = json.load(f)

    feature_names = model_info['feature_names']

    print(f"✓ Modelo cargado: {model_info['model_name']}")
    print(f"✓ Variables: {len(feature_names)}\n")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos para Predicción"

    # ==================== ESTILOS ====================
    border = Border(
        left=Side(style='thin', color='666666'),
        right=Side(style='thin', color='666666'),
        top=Side(style='thin', color='666666'),
        bottom=Side(style='thin', color='666666')
    )

    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')

    pred_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
    pred_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')

    button_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")


    # ==================== ÁREA DEL BOTÓN ====================
    ws.merge_cells('A1:C1')
    ws['A1'] = '🎯 PREDECIR BIOMASA'
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF", name='Arial')
    ws['A1'].fill = button_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Instrucciones del botón
    ws.merge_cells('D1:F1')
    ws['D1'] = '👈 HAZ CLIC EN EL BOTÓN VERDE DESPUÉS DE LLENAR LOS DATOS'
    ws['D1'].font = Font(bold=True, size=10, color="FF6F00", name='Arial')
    ws['D1'].alignment = Alignment(horizontal='left', vertical='center')


    # ==================== INFORMACIÓN ====================
    ws['A2'] = f'📊 Modelo: {model_info["model_name"]} | R²: {model_info["metricas"]["R2_test"]:.4f}'
    ws['A2'].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells('A2:F2')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws['A3'] = ''  # Timestamp se agregará aquí después de predecir
    ws['A3'].font = Font(italic=True, size=9, color="006400")

    ws['A4'] = ''  # Fila vacía


    # ==================== ENCABEZADOS ====================
    header_row = 5

    # ID
    cell = ws.cell(row=header_row, column=1)
    cell.value = 'ID'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

    # Variables
    for col_idx, feature in enumerate(feature_names, start=2):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = feature
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Columna de predicción
    pred_col = len(feature_names) + 2
    cell = ws.cell(row=header_row, column=pred_col)
    cell.value = 'Biomasa_Predicha'
    cell.font = pred_font
    cell.fill = pred_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

    ws.row_dimensions[header_row].height = 30


    # ==================== FILAS DE DATOS ====================
    data_start_row = 6

    for i in range(25):
        row = data_start_row + i

        # ID
        cell = ws.cell(row=row, column=1)
        cell.value = i + 1
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        # Columnas de datos
        for col_idx in range(2, pred_col):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = ""
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Columna de predicción
        cell = ws.cell(row=row, column=pred_col)
        cell.value = ""
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")


    # ==================== AJUSTAR ANCHOS ====================
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6

    for col_idx in range(2, pred_col):
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = 18

    ws.column_dimensions[chr(64 + pred_col)].width = 18


    # ==================== HOJA DE INSTRUCCIONES ====================
    ws_inst = wb.create_sheet("📖 Instrucciones VBA")

    ws_inst['A1'] = '📖 INSTRUCCIONES PARA ACTIVAR EL BOTÓN DE PREDICCIÓN'
    ws_inst['A1'].font = Font(bold=True, size=14, color="D32F2F")
    ws_inst.merge_cells('A1:B1')
    ws_inst['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_inst.row_dimensions[1].height = 25

    instrucciones = [
        "",
        "⚠️ IMPORTANTE: Este archivo necesita configuración VBA para funcionar",
        "━" * 100,
        "",
        "OPCIÓN A: Configuración Manual del Botón (MÁS FÁCIL)",
        "",
        "1. Guarda este archivo como .xlsm (Excel con Macros)",
        "   • File → Save As → Tipo: 'Excel Macro-Enabled Workbook (*.xlsm)'",
        "",
        "2. Habilita la pestaña 'Desarrollador' en Excel",
        "   • File → Options → Customize Ribbon",
        "   • Marca la casilla 'Developer' (Desarrollador)",
        "   • Click OK",
        "",
        "3. Abre el Editor VBA",
        "   • Presiona Alt+F11",
        "   • O ve a Developer → Visual Basic",
        "",
        "4. Inserta un nuevo módulo",
        "   • En el editor VBA: Insert → Module",
        "",
        "5. Copia el código VBA de abajo y pégalo en el módulo",
        "   • Usa el código que está más abajo en esta hoja",
        "",
        "6. Cierra el editor VBA (Alt+Q)",
        "",
        "7. Asigna la macro al botón verde de la hoja principal",
        "   • Ve a la hoja 'Datos para Predicción'",
        "   • Haz clic derecho en la celda verde 'PREDECIR BIOMASA' (A1:C1)",
        "   • Puedes crear un botón: Developer → Insert → Button (Form Control)",
        "   • Arrastra para crear el botón sobre las celdas A1:C1",
        "   • Selecciona la macro 'PredecirBiomasa'",
        "",
        "8. Guarda el archivo y prueba el botón",
        "",
        "",
        "OPCIÓN B: Usar el Script Simplificado (SIN BOTÓN)",
        "",
        "Si no quieres configurar VBA, simplemente:",
        "1. Llena los datos en este Excel",
        "2. Guarda el archivo",
        "3. Ejecuta desde terminal: python predictor_excel_simple.py",
        "",
        "",
        "━" * 100,
        "CÓDIGO VBA PARA COPIAR (después de la línea de puntos)",
        "━" * 100,
        "",
    ]

    row = 1
    for inst in instrucciones:
        ws_inst[f'A{row}'] = inst

        if inst.startswith('━'):
            ws_inst[f'A{row}'].font = Font(color="CCCCCC", size=9)
        elif inst.startswith(('OPCIÓN', 'CÓDIGO VBA')):
            ws_inst[f'A{row}'].font = Font(bold=True, size=12, color="1976D2")
            ws_inst[f'A{row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        elif any(c.isdigit() and '. ' in inst for c in inst[:3]):
            ws_inst[f'A{row}'].font = Font(bold=True, size=10, color="F57C00")
        elif inst.startswith('   •'):
            ws_inst[f'A{row}'].font = Font(size=9, color="555555")
        elif inst.startswith('⚠️'):
            ws_inst[f'A{row}'].font = Font(bold=True, size=11, color="D32F2F")
            ws_inst[f'A{row}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

        row += 1

    # Agregar el código VBA
    vba_lines = crear_codigo_vba().split('\n')
    for vba_line in vba_lines:
        ws_inst[f'A{row}'] = vba_line
        ws_inst[f'A{row}'].font = Font(name='Consolas', size=9, color="000000")
        ws_inst[f'A{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        row += 1

    ws_inst.column_dimensions['A'].width = 120


    # ==================== GUARDAR ====================
    # Guardar como .xlsx primero (el usuario lo convertirá a .xlsm)
    filename = 'Plantilla_Prediccion_Con_Boton.xlsx'
    wb.save(filename)

    # También guardar el código VBA en un archivo separado
    vba_filename = 'codigo_vba_prediccion.bas'
    with open(vba_filename, 'w', encoding='utf-8') as f:
        f.write(crear_codigo_vba())

    print("=" * 80)
    print("✅ EXCEL CON BOTÓN CREADO")
    print("=" * 80)
    print(f"\n📂 Archivos creados:")
    print(f"   1. {filename} - Plantilla Excel")
    print(f"   2. {vba_filename} - Código VBA para el botón")

    print(f"\n📋 Variables incluidas: {len(feature_names)}")
    for i, var in enumerate(feature_names, 1):
        print(f"   {i}. {var}")

    print("\n" + "=" * 80)
    print("PRÓXIMOS PASOS:")
    print("=" * 80)
    print(f"1. Abre: {filename}")
    print("2. Ve a la pestaña '📖 Instrucciones VBA'")
    print("3. Sigue las instrucciones para configurar el botón")
    print("   (O simplemente usa el método sin botón ejecutando predictor_excel_simple.py)")
    print("\n" + "=" * 80 + "\n")

    return filename


if __name__ == "__main__":
    try:
        crear_excel_con_boton()
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
